import mysql.connector
import openpyxl
import os
import csv
import ast

def createDB(passwd, dbname):

	mydb = mysql.connector.connect(
	  host="localhost",
	  user="root",
	  passwd=passwd
	)

	# Get the cursor, which is used to traverse the database, line by line
	mycursor = mydb.cursor()

	mycursor.execute("CREATE DATABASE " + dbname)

	mycursor.execute("SHOW DATABASES")

	for db in mycursor: print(db)

def createTable(passwd, dbname, tblname, strtuple):

	mydb = mysql.connector.connect(
	  host="localhost",
	  user="root",
	  passwd=passwd,
	  database=dbname
	)

	mycursor = mydb.cursor()

	mycursor.execute("CREATE TABLE " + tblname + " " + strtuple)

	mycursor.execute("SHOW TABLES")

	for tbl in mycursor: print(tbl)

def createPrimarykey(passwd, dbname, tblname, strtuple):

	mydb = mysql.connector.connect(
	  host="localhost",
	  user="root",
	  passwd=passwd,
	  database=dbname
	)

	mycursor = mydb.cursor()

	mycursor.execute("ALTER TABLE " + tblname + " ADD PRIMARY KEY " + strtuple)

def createForeignkey(passwd, dbname, tblname, strtuple1, strtuple2):

	mydb = mysql.connector.connect(
	  host="localhost",
	  user="root",
	  passwd=passwd,
	  database=dbname
	)

	mycursor = mydb.cursor()

	mycursor.execute("ALTER TABLE " + tblname + " ADD FOREIGN KEY " + strtuple1 + " REFERENCES " + strtuple2)

def dropTable(passwd, dbname, tblname):

	mydb = mysql.connector.connect(
	  host="localhost",
	  user="root",
	  passwd=passwd,
	  database=dbname
	)

	mycursor = mydb.cursor()

	mycursor.execute("DROP TABLE " + tblname)

def showTables():

	mydb = mysql.connector.connect(
	  host="localhost",
	  user="root",
	  passwd="***",
	  database="SDSS"
	)

	mycursor = mydb.cursor()

	mycursor.execute("SHOW TABLES")

	for tbl in mycursor: print(tbl)

def dataType(val, current_type):
	try:
		# Evaluates numbers to an appropriate type, and strings an error
		t = ast.literal_eval(val)
	except ValueError:
		return "varchar"
	except SyntaxError:
		return "varchar"
	if type(t) is int:
	   # Use smallest possible int type
	   if (int(-32768) < t < int(32767)):
		   return "smallint"
	   elif (int(-2147483648) < t < int(2147483647)):
		   return "int"
	   elif (int(-9223372036854775808) < t < int(9223372036854775807)):
		   return "bigint"
	   else:
		   return "varchar"
	elif type(t) is float:
		return "float"
	else:
		return "varchar"

def scanCSV(path):
	f = open(path, "r")
	reader = csv.reader(f)

	longest, headers, type_list = [], [], []

	for row in reader:
		if len(headers) == 0:
			for col in row:
				headers = row
				longest.append(0)
				type_list.append("")
			n = len(row)
			break

	next(reader)
	for row in reader:
		for i in range(n):
			if len(row[i]) > longest[i]:
				longest[i] = len(row[i])
				var_type = dataType(row[i], type_list[i])
				type_list[i] = var_type
	f.close()

	return longest, headers, type_list

def create_statement(longest, headers, type_list):
	statement = "("
	for i in range(len(headers)):
		if type_list[i] == "varchar":
			statement = (statement + "\n`{}` varchar({}),").format(headers[i].lower(), str(longest[i]))
		else:
			statement = (statement + "\n" + "`{}` {}" + ",").format(headers[i].lower(), type_list[i])
	statement = statement[:-1] + ");\n"

	return statement


if __name__ == "__main__":
	# # Create SDSS DB

	# createDB("***", "SDSS")

	# # Relations creation (our data)

	# createTable("***", "SDSS", "Queries", "(id SMALLINT NOT NULL, timestamp VARCHAR(20) NOT NULL, user VARCHAR(20), database_name VARCHAR(12) NOT NULL, query_text VARCHAR(1500) NOT NULL, running_time FLOAT NOT NULL, output_size INT NOT NULL)")
	# createTable("***", "SDSS", "QueryFeatures", "(query SMALLINT NOT NULL, feature SMALLINT NOT NULL)")
	# createTable("***", "SDSS", "MarginalProbs", "(featureID SMALLINT NOT NULL, probability FLOAT NOT NULL)")
	# createTable("***", "SDSS", "CondProbs", "(feature1 VARCHAR(500) NOT NULL, feature2 VARCHAR(500) NOT NULL, probability FLOAT NOT NULL)")
	# createTable("***", "SDSS", "Features", "(id SMALLINT NOT NULL, feature_description VARCHAR(500) NOT NULL, clause VARCHAR(8) NOT NULL)")

	# # Populate relations (our data)

	# wb = openpyxl.load_workbook("C:/Users/stevi/Desktop/DB/relations/Queries.xlsx")

	# wb = openpyxl.load_workbook("C:/Users/stevi/Desktop/DB/relations/QueryFeatures.xlsx")

	# wb = openpyxl.load_workbook("C:/Users/stevi/Desktop/DB/relations/MarginalProbs.xlsx")

	# wb = openpyxl.load_workbook("C:/Users/stevi/Desktop/DB/relations/Features.xlsx")

	# wb = openpyxl.load_workbook("C:/Users/stevi/Desktop/DB/relations/CondProbs.xlsx")

	# ws = wb["Sheet1"]

	# mydb = mysql.connector.connect(
	#   host="localhost",
	#   user="root",
	#   passwd="***",
	#   database="SDSS"
	# )

	# mycursor = mydb.cursor()

	# # relation Queries
	# sql = "INSERT INTO Queries (id, timestamp, user, database_name, query_text, running_time, output_size) VALUES (%s, %s, %s, %s, %s, %s, %s)"
	# val = [(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value) for row in ws[2: ws.max_row]]

	# # relation QueryFeatures
	# sql = "INSERT INTO QueryFeatures (query, feature) VALUES (%s, %s)"
	# val = [(row[0].value, row[1].value) for row in ws[2: ws.max_row]]

	# # relation MarginalProbs
	# sql = "INSERT INTO MarginalProbs (featureID, probability) VALUES (%s, %s)"
	# val = [(row[0].value, row[1].value) for row in ws[2: ws.max_row]]

	# # relation Features
	# sql = "INSERT INTO Features (id, feature_description, clause) VALUES (%s, %s, %s)"
	# val = [(row[0].value, row[1].value, row[2].value) for row in ws[2: ws.max_row]]

	# # relation CondProbs
	# sql = "INSERT INTO CondProbs (feature1, feature2, probability) VALUES (%s, %s, %s)"
	# val = [(row[0].value, row[1].value, row[2].value) for row in ws[2: ws.max_row]]

	# mycursor.executemany(sql, val)

	# mydb.commit()

	# print("Success!")
	# print(mycursor.rowcount, "was inserted.")

	# # Primary key creation

	# createPrimarykey("***", "SDSS", "Queries", "(id)")
	# createPrimarykey("***", "SDSS", "MarginalProbs", "(featureID)")
	# createPrimarykey("***", "SDSS", "Features", "(id)")

	# # Foreign key creation

	# createForeignkey("***", "SDSS", "QueryFeatures", "(query)", "Queries(id)")
	# createForeignkey("***", "SDSS", "QueryFeatures", "(feature)", "Features(id)")
	# createForeignkey("***", "SDSS", "MarginalProbs", "(featureID)", "QueryFeatures(feature)")
	# createForeignkey("***", "SDSS", "MarginalProbs", "(featureID)", "Features(id)")
	# #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

	# # Relations creation (DR14 data)

	# path = r"C:/ProgramData/MySQL/MySQL Server 8.0/Uploads/SDSS_relations"
	# files = [os.path.splitext(filename)[0] for filename in os.listdir(path)]

	# for f in files:
	# 	path = "C:/ProgramData/MySQL/MySQL Server 8.0/Uploads/SDSS_relations/" + f + ".csv"

	# 	longest, headers, type_list = scanCSV(path)

	# 	statement = create_statement(longest, headers, type_list)

	# 	createTable("***", "SDSS", f, statement)

	# 	print(f)

	# # Primary key creation

	# createPrimarykey("***", "SDSS", "PhotoObjAll", "(objID)")
	# createPrimarykey("***", "SDSS", "Field", "(fieldID)")
	# createPrimarykey("***", "SDSS", "SpecObjAll", "(specObjID)")
	# createPrimarykey("***", "SDSS", "galSpecInfo", "(specObjID)")
	# createPrimarykey("***", "SDSS", "apogeeStar", "(apstar_id)")
	# createPrimarykey("***", "SDSS", "apogeePlate", "(plate_visit_id)")
